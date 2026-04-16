#!/usr/bin/env python3
"""
Servidor webhook simplificado para WhatsApp sem dependências extras.
"""

import asyncio
import logging
import json
import csv
import io
from typing import Dict, Any, List, Optional
from fastapi import FastAPI, Request, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
import uvicorn

from app.message_processor import WhatsAppMessageProcessor, MessageBatchValidator
from config.settings import Settings

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Pydantic models
class MessageItem(BaseModel):
    text: str = Field(..., description="Texto da mensagem")
    delay: Optional[int] = Field(default=2, description="Delay em segundos")

# FastAPI app
app = FastAPI(
    title="WhatsApp Message Dispatcher - Simplified",
    description="Webhook simplificado para envio de mensagens WhatsApp",
    version="1.0.0"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global settings
settings = Settings()
message_processor = WhatsAppMessageProcessor(settings)


@app.get("/")
async def root():
    """Endpoint raiz."""
    return {
        "message": "WhatsApp Message Dispatcher - Simplified",
        "version": "1.0.0",
        "status": "running",
        "endpoints": {
            "webhook": "/webhook/tela2",
            "upload_csv": "/webhook/tela2/upload/csv",
            "upload_xlsx": "/webhook/tela2/upload/xlsx",
            "health": "/health"
        }
    }


@app.get("/health")
async def health_check():
    """Health check."""
    return {"status": "healthy", "timestamp": "2026-04-01"}


@app.post("/webhook/tela2/upload/csv")
async def upload_csv(
    phone_number: str = Form(...),
    file: UploadFile = File(...)
):
    """Upload e processamento de arquivo CSV."""
    try:
        if not file.filename.endswith('.csv'):
            raise HTTPException(status_code=400, detail="Arquivo deve ser CSV")
        
        # Lê o conteúdo do arquivo
        content = await file.read()
        csv_data = content.decode('utf-8')
        
        # Processa o CSV
        messages = process_csv_content(csv_data)
        
        # Envia as mensagens
        results = await message_processor.send_message_batch(
            phone_number=phone_number,
            messages=messages
        )
        
        # Conta sucessos
        sent_count = sum(1 for r in results if r["status"] == "sent")
        
        return {
            "status": "success",
            "message": f"{sent_count} mensagens enviadas com sucesso",
            "total_messages": len(results),
            "sent_messages": sent_count,
            "phone_number": phone_number,
            "results": results
        }
        
    except Exception as e:
        logger.error(f"Erro no upload CSV: {e}")
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/webhook/tela2/upload/xlsx")
async def upload_xlsx(
    phone_number: str = Form(...),
    file: UploadFile = File(...)
):
    """Upload e processamento de arquivo XLSX."""
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Arquivo deve ser XLSX ou XLS")
        
        # Lê o conteúdo do arquivo
        content = await file.read()
        
        # Processa o XLSX (versão simplificada sem pandas)
        messages = process_xlsx_simple(content, file.filename)
        
        # Envia as mensagens
        results = await message_processor.send_message_batch(
            phone_number=phone_number,
            messages=messages
        )
        
        # Conta sucessos
        sent_count = sum(1 for r in results if r["status"] == "sent")
        
        return {
            "status": "success",
            "message": f"{sent_count} mensagens enviadas com sucesso",
            "total_messages": len(results),
            "sent_messages": sent_count,
            "phone_number": phone_number,
            "results": results
        }
        
    except Exception as e:
        logger.error(f"Erro no upload XLSX: {e}")
        raise HTTPException(status_code=400, detail=str(e))
@app.post("/webhook/tela2")
async def webhook_tela2(request: Request):
    """Webhook principal para receber mensagens."""
    try:
        payload = await request.json()
        logger.info(f"Webhook recebido: {payload}")
        
        # Processa o payload
        processed_data = process_webhook_payload(payload)
        
        # Envia as mensagens
        results = await message_processor.send_message_batch(
            phone_number=processed_data["phone_number"],
            messages=processed_data["messages"]
        )
        
        # Conta sucessos
        sent_count = sum(1 for r in results if r["status"] == "sent")
        
        return {
            "status": "success",
            "message": f"{sent_count} mensagens enviadas com sucesso",
            "total_messages": len(results),
            "sent_messages": sent_count,
            "phone_number": processed_data["phone_number"],
            "results": results
        }
        
    except Exception as e:
        logger.error(f"Erro no webhook: {e}")
        raise HTTPException(status_code=400, detail=str(e))


def process_csv_content(csv_content: str) -> List[Dict[str, Any]]:
    """Processa conteúdo CSV."""
    messages = []
    
    # Tenta diferentes delimitadores
    for delimiter in [',', ';', '\t']:
        try:
            csv_reader = csv.DictReader(io.StringIO(csv_content), delimiter=delimiter)
            rows = list(csv_reader)
            
            if rows and len(rows) > 0:
                # Encontra a coluna com texto das mensagens
                text_column = None
                delay_column = None
                
                for col in rows[0].keys():
                    col_lower = col.lower()
                    if any(keyword in col_lower for keyword in ['text', 'message', 'mensagem', 'msg']):
                        text_column = col
                    elif any(keyword in col_lower for keyword in ['delay', 'intervalo', 'tempo']):
                        delay_column = col
                
                # Se não encontrou coluna específica, usa a primeira
                if not text_column:
                    text_column = list(rows[0].keys())[0]
                
                for row in rows:
                    text = row.get(text_column, "").strip()
                    if text:
                        delay = 2
                        if delay_column and row.get(delay_column):
                            try:
                                delay = int(row[delay_column])
                            except:
                                delay = 2
                        
                        messages.append({"text": text, "delay": delay})
                
                break  # Se processou com sucesso, para de tentar outros delimitadores
                
        except Exception as e:
            continue  # Tenta próximo delimitador
    
    if not messages:
        # Fallback: trata como texto simples
        lines = csv_content.strip().split('\n')
        for line in lines:
            line = line.strip()
            if line:
                messages.append({"text": line, "delay": 2})
    
    return messages


def process_xlsx_simple(xlsx_content: bytes, filename: str) -> List[Dict[str, Any]]:
    """Processa XLSX de forma simples."""
    messages = []
    
    try:
        # Tenta usar openpyxl se disponível
        try:
            from openpyxl import load_workbook
            
            # Carrega o workbook do conteúdo
            from io import BytesIO
            workbook = load_workbook(BytesIO(xlsx_content))
            worksheet = workbook.active
            
            # Encontra colunas
            headers = {}
            for col_idx, cell in enumerate(worksheet[1], 1):
                if cell.value:
                    header = str(cell.value).lower()
                    if any(keyword in header for keyword in ['text', 'message', 'mensagem', 'msg']):
                        headers['text'] = col_idx
                    elif any(keyword in header for keyword in ['delay', 'intervalo', 'tempo']):
                        headers['delay'] = col_idx
            
            # Se não encontrou coluna de texto, usa a primeira
            if 'text' not in headers:
                headers['text'] = 1
            
            # Processa linhas (pula header)
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= headers['text']:
                    text = str(row[headers['text'] - 1]).strip() if row[headers['text'] - 1] else ""
                    
                    if text and text != 'None':
                        delay = 2
                        if 'delay' in headers and len(row) >= headers['delay']:
                            try:
                                delay_val = row[headers['delay'] - 1]
                                if delay_val:
                                    delay = int(delay_val)
                            except:
                                delay = 2
                        
                        messages.append({"text": text, "delay": delay})
            
            if messages:
                logger.info(f"✅ Processado XLSX: {len(messages)} mensagens")
                return messages
                
        except ImportError:
            logger.warning("openpyxl não disponível, usando método alternativo")
        
        # Fallback: mensagem explicativa
        messages = [
            {
                "text": f"📄 Arquivo XLSX recebido: {filename}",
                "delay": 2
            },
            {
                "text": "⚠️ Para processar XLSX, instale: pip install openpyxl",
                "delay": 3
            },
            {
                "text": "💡 Alternativa: Salve como CSV ou use 'Colar Lista'",
                "delay": 2
            }
        ]
        
    except Exception as e:
        logger.error(f"Erro ao processar XLSX: {e}")
        messages = [
            {
                "text": f"❌ Erro ao processar {filename}: {str(e)}",
                "delay": 2
            },
            {
                "text": "💡 Tente salvar como CSV e fazer upload novamente",
                "delay": 3
            }
        ]
    
    return messages


def process_webhook_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Processa diferentes tipos de payload."""
    
    # Extrai número de telefone
    phone_number = (
        payload.get("phone_number") or 
        payload.get("chatbotPhone") or
        payload.get("body", {}).get("chatbotPhone")
    )
    
    if not phone_number:
        raise ValueError("Número de telefone é obrigatório")
    
    messages = []
    
    # Formato JSON estruturado
    if "messages" in payload and isinstance(payload["messages"], list):
        for msg in payload["messages"]:
            if isinstance(msg, str):
                messages.append({"text": msg, "delay": 2})
            elif isinstance(msg, dict):
                messages.append({
                    "text": msg.get("text", ""),
                    "delay": msg.get("delay", 2)
                })
    
    # Texto colado
    elif "text_list" in payload:
        lines = payload["text_list"].strip().split('\n')
        for line in lines:
            line = line.strip()
            if line:
                messages.append({"text": line, "delay": 2})
    
    # CSV inline simples
    elif "csv_content" in payload:
        lines = payload["csv_content"].strip().split('\n')
        if len(lines) > 1:  # Tem header
            for line in lines[1:]:  # Pula header
                parts = line.split(',')
                if parts and parts[0].strip():
                    delay = 2
                    if len(parts) > 1:
                        try:
                            delay = int(parts[1].strip())
                        except:
                            delay = 2
                    messages.append({"text": parts[0].strip(), "delay": delay})
        else:
            # Sem header, trata como texto simples
            for line in lines:
                if line.strip():
                    messages.append({"text": line.strip(), "delay": 2})
    
    # Formato legado N8N
    elif "body" in payload and "testCases" in payload["body"]:
        test_cases = payload["body"]["testCases"]
        for case in test_cases:
            messages.append({"text": case, "delay": 2})
    
    # Lista simples de strings
    elif "testCases" in payload:
        for case in payload["testCases"]:
            messages.append({"text": case, "delay": 2})
    
    else:
        raise ValueError("Formato de payload não reconhecido")
    
    # Valida o lote de mensagens
    validation = MessageBatchValidator.validate_message_batch(messages)
    if not validation["valid"]:
        if "errors" in validation:
            raise ValueError(f"Erro na validação: {'; '.join(validation['errors'])}")
        else:
            raise ValueError(f"Erro na validação: {validation['error']}")
    
    return {
        "phone_number": phone_number,
        "messages": messages
    }


if __name__ == "__main__":
    host = settings.WEBHOOK_HOST
    port = settings.WEBHOOK_PORT
    
    logger.info(f"🚀 Iniciando servidor webhook simplificado em {host}:{port}")
    logger.info(f"📱 Webhook URL: http://{host}:{port}/webhook/tela2")
    
    uvicorn.run(
        "app.simple_webhook_server:app",
        host=host,
        port=port,
        reload=True,
        log_level="info"
    )